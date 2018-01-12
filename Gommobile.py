# -*- coding: utf-8 -*-
import time
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

try:
    from itertools import izip
except ImportError:  # python3.x
    izip = zip

#  Acessa os dados de login fora do script, salvo numa planilha existente, para proteger as informações de credenciais
dados = openpyxl.load_workbook('C:\\gomnet.xlsx')
login = dados['Plan1']
url = 'http://gomnet.ampla.com/'
gommobile = 'http://gomnet.ampla.com/Mobile/telaPrincipal.aspx'
username = login['A1'].value
password = login['A2'].value
mobUser = login['B1'].value
mobPass = login['B2'].value
wb = openpyxl.load_workbook('sobs.xlsx')
wb1 = openpyxl.load_workbook('sobs.xlsx')

driver = webdriver.Chrome()

if __name__ == '__main__':
    driver.get(url)
    # Faz login no sistema
    uname = driver.find_element_by_name('txtBoxLogin')
    uname.send_keys(username)
    passw = driver.find_element_by_name('txtBoxSenha')
    passw.send_keys(password)
    submit_button = driver.find_element_by_id('ImageButton_Login').click()
    driver.get(gommobile)

    mobLogin = driver.find_element_by_id('ctl00_ContentPlaceHolder1_txtLogin')
    mobLogin.send_keys(mobUser)
    mobSenha = driver.find_element_by_id('ctl00_ContentPlaceHolder1_txtSenha')
    mobSenha.send_keys(mobPass)
    entrarBtn = driver.find_element_by_id('ctl00_ContentPlaceHolder1_btnEnvia').click()

    # syncBtn = driver.find_element_by_id('ctl00_ContentPlaceHolder1_lbSincronizar').click()
    # time.sleep(5)
    syncBtn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_lbSincronizar')))
    syncBtn.click()
    # okBtn = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/button/span').click()
    okBtn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[3]/div/button/span')))
    okBtn.click()
    # tarefaBtn = driver.find_element_by_id('ctl00_ContentPlaceHolder1_lbTarefaPendente').click()
    tarefaBtn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_lbTarefaPendente')))
    tarefaBtn.click()
    # time.sleep(10)

    for sheet in wb.worksheets:
#        try: # Procura os baremos na planilha "sobs.xlsx" e marca de acordo
            # driver.find_element_by_xpath("//*[contains(text(), '" + str(sheet['A1'].value) + "')]").click()

            obra = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "*//tr/td[contains(text(), '" + str(sheet['A1'].value) + "')]")))
            obra.click()
            # driver.find_element_by_xpath("*//tr/td[contains(text(), '" + str(sheet['A1'].value) + "')]").click()
            #descSob = driver.find_element_by_xpath("*//tbody/tr/td/[contains(text(), 'A019719314')]")
            #webdriver.ActionChains(driver).click(descSob).perform()
            #driver.find_element_by_xpath("//*[@id='gvRelatorio']/tbody/tr/td[contains(text(), '" + str(sheet['A1'].value) + "')]").click()
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform()
            webdriver.ActionChains(driver).send_keys(Keys.RETURN).perform()

            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="gvItens"]/tbody/tr[1]/th[1]'))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tabGeral"]'))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtDtSaida"]'))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[1]'))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="txtIniTarefa"]'))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[1]'))).click()
            # WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/div[3]/div/button[1]/span'))).click() # Finaliza a SOB
            # driver.find_element_by_xpath("*//tr/td[contains(text(), '" + str(cell.value) + "')]/preceding-sibling::td/input").click()
#        except NoSuchElementException:  # Caso não encontre, abre o arquivo txt e registra a data, o código baremo e sua quantidade
        #    log = open("BaremosPendentes.txt", "a")
        #    log.write(str(sheet['H1'].value) + " " + str(sheet['A1'].value) + " " + str(cell.value) + " " + str(cell2.value) + "\n")
        #    log.close()
#           continue
#  Ao fim do loop de inserção de baremos, clica no botão "registrar programação"
#driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnEnviarItens"]').click()
#print(str(sheet['A1'].value) + " programada com êxito.")